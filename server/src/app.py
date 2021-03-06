from flask import Flask, request, redirect, url_for, render_template, jsonify
from flask_pymongo import PyMongo

import os
import json
import glob
import pandas

from openpyxl import Workbook , load_workbook
from openpyxl.utils import column_index_from_string

from config.keys import keys

app = Flask(__name__)

app.config['MONGO_DBNAME'] = keys['MONGO_DBNAME']
app.config['MONGO_URI'] = keys['MONGO_URI']

mongo = PyMongo(app)

def insert_data_db(json_data):
    try:
        converted = mongo.db.converted

        # already have one json with this name
        if (converted.find_one({'FILE_NAME': json_data["FILE_NAME"]})):
            return True
        # insert if do not have
        converted.insert(json_data)

        return True
    except:
        return False

# READ ALL ROWS IN WORKSHEET AND TRANSFORM INTO JSON
def all_data_to_json(worksheet, filename, sheetname):
    max_row = worksheet.max_row
    max_column = worksheet.max_column

    header = []
    for col in worksheet.iter_rows(min_row=1, max_col=max_column, max_row=1):
        for cell in col:
            header.append(cell.value)

    data = []
    for row in worksheet.iter_rows(min_row=2, max_col=max_column, max_row=max_row):
        item = {}
        for cell in row: 
            try:
                if("Date" in header[column_index_from_string(cell.column)-1] or "Data" in header[column_index_from_string(cell.column)-1]):
                    item["Date"] = str(cell.value)
                    continue
            except:
                x = 0
            try:
                item[header[column_index_from_string(cell.column)-1].encode("utf-8")] = cell.value.encode("utf-8")
            except:
                try:
                    item[header[column_index_from_string(cell.column)-1].encode("utf-8")] = cell.value
                except:
                    try:
                        item[header[column_index_from_string(cell.column)-1]] = cell.value.encode("utf-8")
                    except:
                        item['None'] = cell.value
        data.append(item)
    
    json_data = {
        "FILE_NAME" : filename + "_" + sheetname,
        "HEADER" : header,
        "DATA_INFOS" : data,
        "DATA_NUMBER" : max_row - 1
    }

    # insert headers in keys collection
    keys = mongo.db.keys

    for h in header:
        if (keys.find_one({'key': h}) or h == None):
            continue

        # insert if do not have
        keys.insert({'key': h})

    if(insert_data_db(json_data)):
        return json_data["FILE_NAME"]
    

def identifier(data, word):
    for line in data["DATA_INFOS"]:
        for string in line:
            if(string == word):
                return True
        return False

@app.route("/")
def index():
    return "Good job, your flask is working"

@app.route("/api/upload", methods=["POST"])
def upload():
    results = []

    for upload in request.files.getlist("file"):
        # try:
        
        filename = upload.filename.split(".")
        filename = filename[0]

        wb = load_workbook(filename=upload)#, read_only=True)
        sheets = wb.sheetnames
        for sheet in sheets:
            ws = wb[sheet]
            result = all_data_to_json(ws, filename, sheet)
            if result != None:
                results.append(result)
        # except:
        #     print "erro"

    return jsonify(results)

@app.route("/api/download", methods=["GET"])
def download():
    try:
        converted = mongo.db.converted
        json_file = converted.find_one({'FILE_NAME': request.args.get('filename')})

        # delete id form json file
        del json_file["_id"]
        return jsonify(json_file)
    except:
        return jsonify({})

@app.route("/api/search", methods=["GET"])
def search():
    keys = mongo.db.keys

    res = []
    for i in keys.find({"key" : {'$regex': request.args.get('value') , '$options' : 'i'}}):
        res.append(i["key"])
        if len(res) == 10:
            break
    
    return jsonify(res)


@app.route('/api/merge', methods=['GET'])
def merge():
    converted = mongo.db.converted
    
    result_data = []
    
    # get all files in db
    for data in converted.find():
        has_common_identifier = False 
        
        # to each file, check the headers if they have common identifier
        for header in data["HEADER"]:
            if request.args.get('value') == header:
                has_common_identifier = True
                break

        # to each file with the common identifier
        if has_common_identifier == True:
            if result_data:
                for row in data["DATA_INFOS"]:
                    identifier = row[request.args.get('value')]
                    has_updated = False

                    # check if my result data has same identifier, if yes, update with merge
                    for data_row in result_data:
                        if (identifier == data_row[request.args.get('value')]):
                            data_row.update(row)
                            has_updated = True
                            break

                    # if no common identifier was found, add to result without any merge
                    if has_updated == False:
                        result_data.append(row)
            else:
                # result data is empty, insert with the first json found
                for row in data["DATA_INFOS"]:
                    result_data.append(row)

    # json to csv
    df = pandas.read_json(json.dumps(result_data))

    return jsonify(df.to_csv(encoding='utf-8'))
